from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from scripts.build_dashboard_artifacts import clean_current_pick_rows, normalize_pick
from scripts.live_pipeline import load_ballparkpal_snapshot
from tools.ballparkpal.models import DownloadedWorkbook
from tools.ballparkpal.scoring import compute_ballparkpal_overlay
from tools.ballparkpal.snapshot import build_ballparkpal_snapshot
from tools.ballparkpal.validator import validate_workbook_file


class BallparkPalSnapshotAndOverlayTests(unittest.TestCase):
    def _write_workbook(self, path: Path, *, headers: list[str], rows: list[list[object]], sheet_name: str) -> Path:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        workbook.save(path)
        return path

    def test_snapshot_builds_normalized_fields_for_joining(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            base = Path(tmp_dir)
            batters = self._write_workbook(
                base / "batters.xlsx",
                headers=["GameDate", "GamePk", "BatterId", "Team", "Opponent", "HomeRunProbability", "HitProbability"],
                rows=[["2026-04-17", 123, 456, "NYY", "BOS", 0.21, 0.54]],
                sheet_name="Batters",
            )
            pitchers = self._write_workbook(
                base / "pitchers.xlsx",
                headers=["GameDate", "GamePk", "PitcherId", "Team", "Opponent", "RunsAllowed", "HomeRunsAllowed"],
                rows=[["2026-04-17", 123, 789, "BOS", "NYY", 3.1, 0.4]],
                sheet_name="Pitchers",
            )
            teams = self._write_workbook(
                base / "teams.xlsx",
                headers=["GameDate", "GamePk", "Team", "Opponent", "HomeRuns"],
                rows=[["2026-04-17", 123, "NYY", "BOS", 1.2]],
                sheet_name="Teams",
            )
            games = self._write_workbook(
                base / "games.xlsx",
                headers=["GameDate", "GamePk", "HomeTeam", "AwayTeam"],
                rows=[["2026-04-17", 123, "NYY", "BOS"]],
                sheet_name="Games",
            )

            downloads = [
                DownloadedWorkbook("batters", "2026-04-17", "BallparkPal_Batters.xlsx", batters, "https://example.com/batters"),
                DownloadedWorkbook("pitchers", "2026-04-17", "BallparkPal_Pitchers.xlsx", pitchers, "https://example.com/pitchers"),
                DownloadedWorkbook("teams", "2026-04-17", "BallparkPal_Teams.xlsx", teams, "https://example.com/teams"),
                DownloadedWorkbook("games", "2026-04-17", "BallparkPal_Games.xlsx", games, "https://example.com/games"),
            ]
            validations = [
                validate_workbook_file(path, requested_date="2026-04-17", export_name=export_name)
                for export_name, path in [
                    ("batters", batters),
                    ("pitchers", pitchers),
                    ("teams", teams),
                    ("games", games),
                ]
            ]

            snapshot = build_ballparkpal_snapshot(
                requested_date="2026-04-17",
                pulled_at="2026-04-17T12:00:00Z",
                downloads=downloads,
                validations=validations,
            )

            self.assertTrue(snapshot["overall_valid"])
            self.assertEqual(snapshot["requested_date"], "2026-04-17")
            self.assertEqual(snapshot["batters"][0]["home_run_probability"], 0.21)
            self.assertEqual(snapshot["batters"][0]["hit_probability"], 0.54)
            self.assertEqual(snapshot["pitchers"][0]["runs_allowed"], 3.1)
            self.assertEqual(snapshot["pitchers"][0]["home_runs_allowed"], 0.4)
            self.assertEqual(snapshot["teams"][0]["home_runs"], 1.2)

    def test_overlay_scoring_produces_signed_and_display_scores(self) -> None:
        overlay = compute_ballparkpal_overlay(
            {
                "predicted_hr_score": 82.0,
                "ballparkpal_home_run_probability": 0.23,
            }
        )

        self.assertEqual(overlay["ballparkpal_overlay_signed_score"], 13.0)
        self.assertEqual(overlay["ballparkpal_overlay_display_score"], 23.0)
        self.assertIsNotNone(overlay["ballparkpal_overlay_adjusted_score"])
        self.assertEqual(overlay["ballparkpal_overlay_blend_weights"]["model"], 0.1)
        self.assertEqual(overlay["ballparkpal_overlay_blend_weights"]["ballpark"], 0.9)
        expected_adjusted = round(
            0.1 * overlay["ballparkpal_overlay_model_score"] + 0.9 * overlay["ballparkpal_overlay_display_score"],
            1,
        )
        self.assertEqual(overlay["ballparkpal_overlay_adjusted_score"], expected_adjusted)
        self.assertEqual(overlay["ballparkpal_overlay_direction"], "favorable")
        self.assertIn("ballparkpal_home_run_probability", overlay["ballparkpal_overlay_factor_details"])
        self.assertEqual(
            set(overlay["ballparkpal_overlay_factor_details"].keys()),
            {"ballparkpal_home_run_probability"},
        )

    def test_overlay_scoring_without_ballpark_data_falls_back_to_model_score(self) -> None:
        overlay = compute_ballparkpal_overlay({"predicted_hr_score": 82.0})

        self.assertEqual(overlay["ballparkpal_overlay_signed_score"], 0.0)
        self.assertEqual(overlay["ballparkpal_overlay_display_score"], 50.0)
        self.assertEqual(overlay["ballparkpal_overlay_adjusted_score"], 82.0)
        self.assertEqual(overlay["ballparkpal_overlay_direction"], "neutral")

    def test_dashboard_cleaning_preserves_ballparkpal_columns(self) -> None:
        normalized = normalize_pick(
            {
                "game_date": "2026-04-17",
                "game_pk": 123,
                "batter_id": 456,
                "batter_name": "Alpha",
                "team": "NYY",
                "opponent_team": "BOS",
                "pitcher_id": 789,
                "pitcher_name": "Pitcher A",
                "predicted_hr_probability": 0.23,
                "predicted_hr_score": 82.0,
                "ballparkpal_snapshot_status": "loaded",
                "ballparkpal_snapshot_date": "2026-04-17",
                "ballparkpal_team_home_runs": 1.42,
                "ballparkpal_overlay_display_score": 73.5,
                "ballparkpal_overlay_direction": "favorable",
                "result": "Pending",
            },
            "2026-04-17",
        )
        cleaned = clean_current_pick_rows([normalized])

        self.assertEqual(cleaned[0]["ballparkpal_snapshot_status"], "loaded")
        self.assertEqual(cleaned[0]["ballparkpal_team_home_runs"], 1.42)
        self.assertEqual(cleaned[0]["ballparkpal_overlay_display_score"], 73.5)
        self.assertEqual(cleaned[0]["ballparkpal_overlay_direction"], "favorable")

    def test_history_cleaning_preserves_ballparkpal_columns(self) -> None:
        from scripts.build_dashboard_artifacts import clean_history_rows

        normalized = normalize_pick(
            {
                "game_date": "2026-04-17",
                "game_pk": 123,
                "batter_id": 456,
                "batter_name": "Alpha",
                "team": "NYY",
                "opponent_team": "BOS",
                "pitcher_id": 789,
                "pitcher_name": "Pitcher A",
                "predicted_hr_probability": 0.23,
                "predicted_hr_score": 82.0,
                "ballparkpal_snapshot_status": "loaded",
                "ballparkpal_snapshot_date": "2026-04-17",
                "ballparkpal_team_home_runs": 1.42,
                "ballparkpal_overlay_display_score": 73.5,
                "ballparkpal_overlay_direction": "favorable",
                "result": "HR",
            },
            "2026-04-17",
        )
        cleaned = clean_history_rows([normalized])

        self.assertEqual(cleaned[0]["ballparkpal_snapshot_status"], "loaded")
        self.assertEqual(cleaned[0]["ballparkpal_team_home_runs"], 1.42)
        self.assertEqual(cleaned[0]["ballparkpal_overlay_display_score"], 73.5)
        self.assertEqual(cleaned[0]["ballparkpal_overlay_direction"], "favorable")

    def test_stale_latest_snapshot_is_ignored_when_requested_date_differs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            base = Path(tmp_dir)
            latest_snapshot = base / "latest_snapshot.json"
            latest_snapshot.write_text(
                json.dumps(
                    {
                        "requested_date": "2026-04-17",
                        "pulled_at": "2026-04-17T06:10:00Z",
                        "overall_valid": True,
                    }
                ),
                encoding="utf-8",
            )

            with patch("scripts.live_pipeline.BALLPARKPAL_LATEST_SNAPSHOT_PATH", latest_snapshot), patch(
                "scripts.live_pipeline.BALLPARKPAL_VALIDATED_DIR", base / "validated"
            ):
                snapshot = load_ballparkpal_snapshot("2026-04-18")

        self.assertIsNone(snapshot)
