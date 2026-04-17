"""Optional validation overlay helpers built from validated Ballpark Pal exports."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .validator import _normalize_date_string


def _read_first_matching_sheet(path: Path, hints: tuple[str, ...]) -> pd.DataFrame:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            normalized = sheet_name.lower()
            if not any(hint in normalized for hint in hints):
                continue
            worksheet = workbook[sheet_name]
            rows = list(worksheet.values)
            if not rows:
                continue
            headers = [str(value).strip() if value not in (None, "") else f"column_{index + 1}" for index, value in enumerate(rows[0])]
            return pd.DataFrame(rows[1:], columns=headers)
        return pd.DataFrame()
    finally:
        workbook.close()


def _pick_first_present(row: pd.Series, candidates: tuple[str, ...]) -> Any:
    for candidate in candidates:
        if candidate in row and row[candidate] not in (None, ""):
            return row[candidate]
    return None


def _sheet_overlay(frame: pd.DataFrame, *, requested_date: str, name_candidates: tuple[str, ...]) -> list[dict[str, Any]]:
    if frame.empty:
        return []
    normalized_requested = _normalize_date_string(requested_date)
    overlay_rows: list[dict[str, Any]] = []
    for _, row in frame.iterrows():
        row_date = _pick_first_present(row, ("GameDate", "SlateDate", "Date", "game_date", "date"))
        normalized_row_date = _normalize_date_string(row_date)
        if normalized_requested is not None and normalized_row_date not in (None, normalized_requested):
            continue
        overlay_rows.append(
            {
                "name": _pick_first_present(row, name_candidates),
                "game_date": normalized_row_date,
                "home_run_probability": _pick_first_present(row, ("HomeRunProbability", "home_run_probability")),
                "hit_probability": _pick_first_present(row, ("HitProbability", "hit_probability")),
                "runs_allowed": _pick_first_present(row, ("RunsAllowed", "runs_allowed")),
                "home_runs_allowed": _pick_first_present(row, ("HomeRunsAllowed", "home_runs_allowed")),
            }
        )
    return overlay_rows


def build_validation_overlay(paths_by_export: dict[str, Path], requested_date: str) -> dict[str, Any]:
    payload = {
        "requested_date": requested_date,
        "batters": _sheet_overlay(
            _read_first_matching_sheet(paths_by_export["batters"], ("batter", "batters")),
            requested_date=requested_date,
            name_candidates=("Batter", "Player", "Name"),
        ),
        "pitchers": _sheet_overlay(
            _read_first_matching_sheet(paths_by_export["pitchers"], ("pitcher", "pitchers")),
            requested_date=requested_date,
            name_candidates=("Pitcher", "Player", "Name"),
        ),
        "teams": _sheet_overlay(
            _read_first_matching_sheet(paths_by_export["teams"], ("team", "teams")),
            requested_date=requested_date,
            name_candidates=("Team", "Name"),
        ),
        "games": _sheet_overlay(
            _read_first_matching_sheet(paths_by_export["games"], ("game", "games")),
            requested_date=requested_date,
            name_candidates=("Game", "Matchup", "Name"),
        ),
    }
    return payload


def write_validation_overlay(payload: dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    return path
