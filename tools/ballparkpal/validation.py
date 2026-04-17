"""Validation helpers for Ballpark Pal export files."""

from __future__ import annotations

import hashlib
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


# Tolerant minimum expected columns (substring matching, case-insensitive).
EXPECTED_COLUMN_HINTS: dict[str, tuple[str, ...]] = {
    "batters": ("player", "team", "opp", "prob"),
    "pitchers": ("pitcher", "team", "opp", "proj"),
    "teams": ("team", "opp", "proj"),
    "games": ("away", "home", "proj"),
}


@dataclass
class ValidationResult:
    is_valid: bool
    error: str | None
    exists: bool
    extension_ok: bool
    binary_signature_ok: bool
    workbook_open_ok: bool
    detected_sheets: list[str]
    detected_columns: list[str]
    expected_columns_matched: int
    expected_columns_total: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def sha256sum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _looks_like_html_or_text(path: Path) -> bool:
    head = path.read_bytes()[:2048]
    lowered = head.lower()
    if lowered.startswith(b"pk\x03\x04"):
        return False
    html_signatures = (b"<!doctype html", b"<html", b"<head", b"<body", b"<title", b"<?xml")
    if any(sig in lowered for sig in html_signatures):
        return True
    if b"login" in lowered and b"password" in lowered:
        return True
    # Plain text files often have no NUL bytes in the first chunk.
    if b"\x00" not in head and b"," in head:
        return True
    return False


def _normalize_columns(columns: list[str]) -> list[str]:
    normalized: list[str] = []
    for col in columns:
        text = str(col).strip().lower()
        if text and text != "nan":
            normalized.append(text)
    return normalized


def validate_export_file(path: Path, export_key: str) -> ValidationResult:
    exists = path.exists()
    extension_ok = path.suffix.lower() == ".xlsx"
    if not exists:
        return ValidationResult(
            is_valid=False,
            error="File does not exist.",
            exists=False,
            extension_ok=extension_ok,
            binary_signature_ok=False,
            workbook_open_ok=False,
            detected_sheets=[],
            detected_columns=[],
            expected_columns_matched=0,
            expected_columns_total=len(EXPECTED_COLUMN_HINTS.get(export_key, ())),
        )

    binary_signature_ok = not _looks_like_html_or_text(path)
    if not extension_ok:
        return ValidationResult(
            is_valid=False,
            error=f"Expected .xlsx extension, found {path.suffix}.",
            exists=True,
            extension_ok=False,
            binary_signature_ok=binary_signature_ok,
            workbook_open_ok=False,
            detected_sheets=[],
            detected_columns=[],
            expected_columns_matched=0,
            expected_columns_total=len(EXPECTED_COLUMN_HINTS.get(export_key, ())),
        )

    if not binary_signature_ok:
        return ValidationResult(
            is_valid=False,
            error="File signature looks like HTML/text and not a valid XLSX binary.",
            exists=True,
            extension_ok=True,
            binary_signature_ok=False,
            workbook_open_ok=False,
            detected_sheets=[],
            detected_columns=[],
            expected_columns_matched=0,
            expected_columns_total=len(EXPECTED_COLUMN_HINTS.get(export_key, ())),
        )

    detected_sheets: list[str] = []
    workbook_open_ok = False
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        detected_sheets = [str(name) for name in workbook.sheetnames]
        workbook_open_ok = len(detected_sheets) > 0
        workbook.close()
    except Exception as exc:
        return ValidationResult(
            is_valid=False,
            error=f"Workbook open failed: {exc}",
            exists=True,
            extension_ok=True,
            binary_signature_ok=True,
            workbook_open_ok=False,
            detected_sheets=[],
            detected_columns=[],
            expected_columns_matched=0,
            expected_columns_total=len(EXPECTED_COLUMN_HINTS.get(export_key, ())),
        )

    detected_columns: list[str] = []
    try:
        df = pd.read_excel(path, sheet_name=0, nrows=10)
        detected_columns = _normalize_columns(list(df.columns))
    except Exception as exc:
        return ValidationResult(
            is_valid=False,
            error=f"Header read failed: {exc}",
            exists=True,
            extension_ok=True,
            binary_signature_ok=True,
            workbook_open_ok=workbook_open_ok,
            detected_sheets=detected_sheets,
            detected_columns=[],
            expected_columns_matched=0,
            expected_columns_total=len(EXPECTED_COLUMN_HINTS.get(export_key, ())),
        )

    hints = EXPECTED_COLUMN_HINTS.get(export_key, ())
    matched = 0
    for hint in hints:
        if any(hint in column for column in detected_columns):
            matched += 1

    # Tolerant schema check: at least half of expected hint tokens must appear.
    minimum_match = max(1, len(hints) // 2) if hints else 0
    schema_ok = matched >= minimum_match
    is_valid = exists and extension_ok and binary_signature_ok and workbook_open_ok and schema_ok
    error: str | None = None
    if not schema_ok:
        error = (
            f"Schema mismatch: matched {matched}/{len(hints)} expected column hints "
            f"for export {export_key}."
        )

    return ValidationResult(
        is_valid=is_valid,
        error=error,
        exists=exists,
        extension_ok=extension_ok,
        binary_signature_ok=binary_signature_ok,
        workbook_open_ok=workbook_open_ok,
        detected_sheets=detected_sheets,
        detected_columns=detected_columns,
        expected_columns_matched=matched,
        expected_columns_total=len(hints),
    )

