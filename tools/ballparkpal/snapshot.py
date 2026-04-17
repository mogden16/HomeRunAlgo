"""Normalize validated Ballpark Pal export workbooks into join-ready JSON snapshots."""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import DownloadedWorkbook, ValidationFinding

EXPORT_SHEET_HINTS: dict[str, tuple[str, ...]] = {
    "batters": ("batter", "batters"),
    "pitchers": ("pitcher", "pitchers"),
    "teams": ("team", "teams"),
    "games": ("game", "games"),
}


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _coerce_scalar(value: object) -> object:
    if value is None:
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:  # pragma: no cover - numpy/pandas compatibility
            return value
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc).isoformat()
        return value.astimezone(timezone.utc).isoformat()
    return value


def _normalize_date(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    for candidate in (text, text.replace("Z", "+00:00"), text.split("T", 1)[0], text.split(" ", 1)[0]):
        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError:
            continue
        return parsed.date().isoformat()
    return text


def _normalize_number(value: object) -> object:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    try:
        number = float(str(value))
    except (TypeError, ValueError):
        return value
    return int(number) if number.is_integer() else number


def _sheet_matches(sheet_name: str, hints: tuple[str, ...]) -> bool:
    normalized = sheet_name.lower()
    return any(hint in normalized for hint in hints)


def _find_header_row(rows: list[tuple[object, ...]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:10]):
        non_empty = [cell for cell in row if cell not in (None, "")]
        if not non_empty:
            continue
        score = len(non_empty)
        for cell in non_empty:
            text = str(cell).lower()
            if any(token in text for token in ("date", "game", "team", "player", "batter", "pitcher", "probability", "allowed")):
                score += 2
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _build_header_map(headers: list[object]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_text(header)
        if normalized and normalized not in header_map:
            header_map[normalized] = index
    return header_map


def _get_first_present(row: list[object], header_map: dict[str, int], aliases: tuple[str, ...]) -> object | None:
    for alias in aliases:
        index = header_map.get(_normalize_text(alias))
        if index is None or index >= len(row):
            continue
        value = row[index]
        if value not in (None, ""):
            return value
    return None


def _extract_rows(path: Path, *, export_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_names = [
            sheet_name
            for sheet_name in workbook.sheetnames
            if _sheet_matches(sheet_name, EXPORT_SHEET_HINTS[export_name])
        ]
        if not sheet_names:
            sheet_names = list(workbook.sheetnames)
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
            if not rows:
                continue
            header_index = _find_header_row(rows)
            headers = list(rows[header_index])
            header_map = _build_header_map(headers)
            records: list[dict[str, Any]] = []
            for raw_row in rows[header_index + 1 :]:
                if all(cell in (None, "") for cell in raw_row):
                    continue
                normalized_row: dict[str, Any] = {}
                for index, header in enumerate(headers):
                    if index >= len(raw_row):
                        continue
                    key = _normalize_text(header) or f"column_{index + 1}"
                    normalized_row[key] = _coerce_scalar(raw_row[index])
                if export_name == "batters":
                    normalized_row["game_date"] = _normalize_date(
                        _get_first_present(raw_row, header_map, ("game_date", "gamedate", "slate_date", "date", "contest_date", "export_date"))
                    )
                    normalized_row["game_pk"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("game_pk", "gamepk", "game", "matchup_id"))
                    )
                    normalized_row["batter_id"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("batter_id", "player_id", "playerid", "player", "id"))
                    )
                    normalized_row["team"] = str(
                        _get_first_present(raw_row, header_map, ("team", "batting_team", "club"))
                        or ""
                    )
                    normalized_row["opponent"] = str(
                        _get_first_present(raw_row, header_map, ("opponent", "opp_team", "opposing_team"))
                        or ""
                    )
                    normalized_row["home_run_probability"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("home_run_probability", "homerunprobability", "hr_probability"))
                    )
                    normalized_row["hit_probability"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("hit_probability", "hitprobability"))
                    )
                elif export_name == "pitchers":
                    normalized_row["game_date"] = _normalize_date(
                        _get_first_present(raw_row, header_map, ("game_date", "gamedate", "slate_date", "date", "contest_date", "export_date"))
                    )
                    normalized_row["game_pk"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("game_pk", "gamepk", "game", "matchup_id"))
                    )
                    normalized_row["pitcher_id"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("pitcher_id", "player_id", "playerid", "player", "id"))
                    )
                    normalized_row["team"] = str(
                        _get_first_present(raw_row, header_map, ("team", "pitching_team", "club"))
                        or ""
                    )
                    normalized_row["opponent"] = str(
                        _get_first_present(raw_row, header_map, ("opponent", "opp_team", "batting_team"))
                        or ""
                    )
                    normalized_row["runs_allowed"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("runs_allowed", "runsallowed", "run_allowed", "allowed_runs"))
                    )
                    normalized_row["home_runs_allowed"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("home_runs_allowed", "homeruns_allowed", "homerunsallowed", "hr_allowed"))
                    )
                elif export_name == "teams":
                    normalized_row["game_date"] = _normalize_date(
                        _get_first_present(raw_row, header_map, ("game_date", "gamedate", "slate_date", "date", "contest_date", "export_date"))
                    )
                    normalized_row["game_pk"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("game_pk", "gamepk", "game", "matchup_id"))
                    )
                    normalized_row["team"] = str(_get_first_present(raw_row, header_map, ("team", "club")) or "")
                    normalized_row["opponent"] = str(_get_first_present(raw_row, header_map, ("opponent", "opp_team")) or "")
                    normalized_row["home_runs"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("home_runs", "homeruns", "hr", "home_run"))
                    )
                elif export_name == "games":
                    normalized_row["game_date"] = _normalize_date(
                        _get_first_present(raw_row, header_map, ("game_date", "gamedate", "slate_date", "date", "contest_date", "export_date"))
                    )
                    normalized_row["game_pk"] = _normalize_number(
                        _get_first_present(raw_row, header_map, ("game_pk", "gamepk", "game", "matchup_id"))
                    )
                    normalized_row["home_team"] = str(_get_first_present(raw_row, header_map, ("home_team", "home", "home club")) or "")
                    normalized_row["away_team"] = str(_get_first_present(raw_row, header_map, ("away_team", "away", "visitor", "visiting_team")) or "")
                records.append(normalized_row)
            return records
        return []
    finally:
        workbook.close()


def build_ballparkpal_snapshot(
    *,
    requested_date: str,
    pulled_at: str,
    downloads: list[DownloadedWorkbook],
    validations: list[ValidationFinding],
) -> dict[str, Any]:
    downloads_by_export = {item.export_name: item for item in downloads}
    validations_by_export = {item.export_name: item for item in validations}
    snapshot: dict[str, Any] = {
        "requested_date": requested_date,
        "pulled_at": pulled_at,
        "overall_valid": all(item.valid for item in validations),
        "exports": {
            export_name: {
                "requested_date": requested_date,
                **validations_by_export[export_name].as_manifest_row(),
                "source_url": downloads_by_export[export_name].source_url,
                "original_filename": downloads_by_export[export_name].original_filename,
            }
            for export_name in downloads_by_export
            if export_name in validations_by_export
        },
    }
    for export_name in ("batters", "pitchers", "teams", "games"):
        download = downloads_by_export.get(export_name)
        if download is None:
            snapshot[export_name] = []
            continue
        records = _extract_rows(download.saved_path, export_name=export_name)
        filtered: list[dict[str, Any]] = []
        for record in records:
            record_date = _normalize_date(record.get("game_date"))
            if record_date not in (None, requested_date):
                continue
            filtered.append(record)
        snapshot[export_name] = filtered
    return snapshot


def write_ballparkpal_snapshot(snapshot: dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(snapshot, indent=2, sort_keys=True), encoding="utf-8")
    return path


def write_latest_ballparkpal_snapshot(snapshot: dict[str, Any], path: Path) -> Path:
    return write_ballparkpal_snapshot(snapshot, path)
