"""Validation helpers for Ballpark Pal workbook exports."""

from __future__ import annotations

import hashlib
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

from .models import ValidationFinding


DATE_HEADER_KEYWORDS = ("date", "slate", "game date", "game_date", "contest date", "export date")
IDENTIFIER_HEADER_KEYWORDS = ("player", "batter", "pitcher", "team", "game", "name")


@dataclass(frozen=True)
class WorkbookExpectation:
    sheet_name_hints: tuple[str, ...]
    required_headers: tuple[str, ...]
    date_header_keywords: tuple[str, ...]
    required_row_count: int = 1


EXPORT_EXPECTATIONS: dict[str, WorkbookExpectation] = {
    "batters": WorkbookExpectation(
        sheet_name_hints=("batter", "batters"),
        required_headers=("HomeRunProbability", "HitProbability"),
        date_header_keywords=("date", "game date", "slate date"),
    ),
    "pitchers": WorkbookExpectation(
        sheet_name_hints=("pitcher", "pitchers"),
        required_headers=("RunsAllowed", "HomeRunsAllowed"),
        date_header_keywords=("date", "game date", "slate date"),
    ),
    "teams": WorkbookExpectation(
        sheet_name_hints=("team", "teams"),
        required_headers=("Team",),
        date_header_keywords=("date", "game date", "slate date"),
    ),
    "games": WorkbookExpectation(
        sheet_name_hints=("game", "games"),
        required_headers=("HomeTeam", "AwayTeam"),
        date_header_keywords=("date", "game date", "slate date"),
    ),
}


def _normalize_text(value: object) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _normalize_date_string(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for candidate in (text, text.replace("Z", "+00:00"), text.split("T", 1)[0], text.split(" ", 1)[0]):
        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError:
            continue
        return parsed.date().isoformat()
    return None


def _looks_like_date_header(header: str, keywords: tuple[str, ...]) -> bool:
    normalized = header.strip().lower()
    return any(keyword in normalized for keyword in keywords)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_header_and_rows(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[str], list[list[object]], int]:
    best_row_index = 1
    best_score = -1
    max_scan = min(10, worksheet.max_row or 1)
    for row_index in range(1, max_scan + 1):
        values = [cell.value for cell in next(worksheet.iter_rows(min_row=row_index, max_row=row_index))]
        non_empty = [value for value in values if value not in (None, "")]
        if not non_empty:
            continue
        score = len(non_empty)
        score += sum(
            2 for value in non_empty if isinstance(value, str) and any(token in value.lower() for token in IDENTIFIER_HEADER_KEYWORDS + DATE_HEADER_KEYWORDS)
        )
        if score > best_score:
            best_score = score
            best_row_index = row_index
    header_values = [cell.value for cell in next(worksheet.iter_rows(min_row=best_row_index, max_row=best_row_index))]
    headers = [str(value).strip() if value not in (None, "") else f"column_{index + 1}" for index, value in enumerate(header_values)]
    rows: list[list[object]] = []
    for row in worksheet.iter_rows(min_row=best_row_index + 1, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        rows.append(list(row))
    return headers, rows, best_row_index


def _match_required_headers(headers: list[str], required_headers: tuple[str, ...]) -> list[str]:
    normalized_headers = {_normalize_text(header) for header in headers}
    missing = []
    for required in required_headers:
        required_normalized = _normalize_text(required)
        if required_normalized not in normalized_headers:
            missing.append(required)
    return missing


def _sheet_matches_hint(name: str, hints: tuple[str, ...]) -> bool:
    normalized = name.lower()
    return any(hint in normalized for hint in hints)


def _candidate_date_values(headers: list[str], rows: list[list[object]], keywords: tuple[str, ...]) -> list[str]:
    date_columns = [index for index, header in enumerate(headers) if _looks_like_date_header(header, keywords)]
    candidates: list[str] = []
    for row in rows:
        for index in date_columns:
            if index >= len(row):
                continue
            normalized = _normalize_date_string(row[index])
            if normalized is not None:
                candidates.append(normalized)
    if candidates:
        return candidates

    for row in rows:
        for cell in row:
            normalized = _normalize_date_string(cell)
            if normalized is not None:
                candidates.append(normalized)
    return candidates


def validate_workbook_file(
    path: Path,
    *,
    requested_date: str,
    export_name: str,
) -> ValidationFinding:
    errors: list[str] = []
    warnings: list[str] = []
    if not path.exists():
        return ValidationFinding(
            valid=False,
            export_name=export_name,
            requested_date=requested_date,
            saved_path=path,
            errors=("file does not exist",),
        )
    if path.suffix.lower() != ".xlsx":
        errors.append("file does not use the .xlsx extension")

    file_size_bytes = path.stat().st_size
    if file_size_bytes <= 0:
        errors.append("file is empty")

    raw_prefix = path.read_bytes()[:512].lstrip()
    lower_prefix = raw_prefix.lower()
    if lower_prefix.startswith(b"<html") or lower_prefix.startswith(b"<!doctype html") or b"<html" in lower_prefix[:128]:
        errors.append("file looks like HTML, not an Excel workbook")
    if not zipfile.is_zipfile(path):
        errors.append("file is not a valid .xlsx zip archive")

    if errors:
        return ValidationFinding(
            valid=False,
            export_name=export_name,
            requested_date=requested_date,
            saved_path=path,
            errors=tuple(errors),
            warnings=tuple(warnings),
            sha256=_sha256(path) if path.exists() and file_size_bytes > 0 else None,
            file_size_bytes=file_size_bytes,
        )

    workbook = None
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        expectation = EXPORT_EXPECTATIONS.get(export_name)
        sheet_names = tuple(workbook.sheetnames)
        if not sheet_names:
            errors.append("workbook has no sheets")
        if expectation is not None and not any(_sheet_matches_hint(name, expectation.sheet_name_hints) for name in sheet_names):
            warnings.append("no sheet name matched the expected export hint")

        workbook_dates: set[str] = set()
        total_rows = 0
        for worksheet in workbook.worksheets:
            headers, rows, _header_row_index = _read_header_and_rows(worksheet)
            total_rows += len(rows)
            if expectation is not None:
                missing_headers = _match_required_headers(headers, expectation.required_headers)
                if missing_headers:
                    errors.append(f"sheet '{worksheet.title}' is missing required headers: {', '.join(missing_headers)}")
            candidates = _candidate_date_values(
                headers,
                rows,
                expectation.date_header_keywords if expectation is not None else DATE_HEADER_KEYWORDS,
            )
            workbook_dates.update(candidate for candidate in candidates if candidate is not None)
    except Exception as exc:  # pragma: no cover - converted to validation failure
        errors.append(f"openpyxl could not open workbook: {exc}")
        return ValidationFinding(
            valid=False,
            export_name=export_name,
            requested_date=requested_date,
            saved_path=path,
            errors=tuple(errors),
            warnings=tuple(warnings),
            sha256=_sha256(path),
            file_size_bytes=file_size_bytes,
        )
    finally:
        if workbook is not None:
            workbook.close()

    normalized_requested_date = _normalize_date_string(requested_date)
    if normalized_requested_date is None:
        errors.append("requested date is not a valid date")
    if not workbook_dates:
        errors.append("no workbook date fields were found")
    elif len(workbook_dates) > 1:
        errors.append(f"workbook contains mismatched dates: {sorted(workbook_dates)}")
    elif normalized_requested_date is not None and next(iter(workbook_dates)) != normalized_requested_date:
        errors.append(f"workbook date {next(iter(workbook_dates))} does not match requested date {normalized_requested_date}")

    if expectation is not None and total_rows < expectation.required_row_count:
        errors.append(f"workbook contains fewer than {expectation.required_row_count} data rows")

    return ValidationFinding(
        valid=not errors,
        export_name=export_name,
        requested_date=requested_date,
        saved_path=path,
        workbook_date=next(iter(workbook_dates)) if len(workbook_dates) == 1 else None,
        sheet_names=sheet_names,
        row_count=total_rows,
        errors=tuple(errors),
        warnings=tuple(warnings),
        sha256=_sha256(path),
        file_size_bytes=file_size_bytes,
    )
